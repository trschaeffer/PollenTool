# -*- coding: utf-8 -*-
"""
Created on Tue Mar 17 16:37:06 2020

@author: Tobias
"""

from openpyxl import Workbook,load_workbook
from datetime import date as d
from datetime import datetime as dt
import xlrd
import numpy as np
from pollenCategories import pollenCategories
months= ['January', 'February', 'March', 'April', 'May', 'June', 'July', 
         'August', 'September', 'October', 'November', 'December'] 
#reads file and returns worksheet
def readFile(fileName):
    if(fileName.split('.')[-1])=='xlsx':
        wb = load_workbook(fileName)
    #grab the active worksheet
        ws = wb.active
    elif(fileName.split('.')[-1])=='xls':
        wb = xlrd.open_workbook(fileName)
        ws = wb.sheet_by_index(0)
    return ws
#finds the year of the spreadsheet, used if the spreadsheet is in month format
def findYear(ws):
    year=0
    #year is specified in the first cell in the spread sheet
    try:
        cell = ws['A1'].value
    except TypeError:
        cell= ws.cell(0,0).value
    if cell!=None:
        words=cell.split(' ') #separate words
    else:
        words=["a"]
    #extract the year from the first cell
    for word in words:
        try:
            if(int(word)>1900 and int(word)<2100):
                year=int(word)
                break
            
        except ValueError:
            try:
                
                orientation=findOrientation(ws)
                
                if(orientation==0):
                    
                    year=ws['B1'].value.year
                elif(orientation==1):
                    year=ws['A2'].value.year
                elif(orientation==2):
                    year=xlrd.xldate_as_tuple(ws.cell(2,0).value,0)[0]
                    
            except Exception:
                pass
           
    return year

#determine where the dates are formatted in the spreadsheet
    #1=dates are beside rows
#0=dates are above columns, categories on rows
#2=data comes from xls, dates on left, starting at cell 2
    #-1=error
def findOrientation(ws):
    try:
        if(dateType(ws['A2'].value)>0): 
            orientation=1
        elif(dateType(ws['B1'].value)>0): 
            orientation=0
        else:
            return -1
    except:
        date=xlrd.xldate_as_tuple(ws.cell(2,0).value,0)
        date=(d(date[0],date[1],date[2]))
        if(dateType(date)>0):
            orientation= 2
     
    
    return orientation

#turns excel worksheet into a data set
#after this, the data set should be filtered to remove empty categories and dates
#using filter()

def toList(ws,orientation,isP):
    year = findYear(ws)
    
    #extract all data from rows/columns depending on orientation
    if(orientation==0):
        alldata=ws.rows
    elif(orientation==1):
        alldata=ws.columns
    elif(orientation==2):
        dates=[]
        categories=[]
        data=[]
        #special code for .xls files, cannot be reused/ mixed with .xlsx
        for i in range(1,ws.nrows):
            if(i>1):
                date=xlrd.xldate_as_tuple(ws.cell(i,0).value,0)
                dates.append(d(date[0],date[1],date[2]))
            temparray=[]    
            for j in range(1, ws.ncols):
                if(i==1):
                    categories.append(ws.cell(i,j).value)
                    
                else:
                    if(ws.cell(i,j).value!=''):
                        temparray.append(ws.cell(i,j).value)
                    else:
                        temparray.append(None)
            data.append(temparray)
        del data[0]
        data=list(map(list, zip(*data)))
        
        
        return categories,data,dates
    #.xlsx code which can handle multiple different format
    categories=[]
    listdata=[]
    alldata=(tuple(alldata))
    
    #separate data from categories
    for row in tuple(alldata)[1:]:
        data=[]
        for cell in row[1:]:
            data.append(cell.value)
        #print(data)
        if(row[0].value!=None):
            #if the user wants the (p) in the file, set isP to true
            #typically used in backend functions
            if(isP):
                categories.append(row[0].value)
            else:
            #otherwise the (p) is filtered out so the user cannot see it. 
            #typically used in frontend functions
                categories.append(row[0].value.split('(p')[0])
            listdata.append(data.copy())
    #establish dates list        
    dates=[]
    daterow=alldata[0][1:]
    
    #create a list of all the dates
    for date in daterow:
        if(date.value!=None):
            #if date is written out as months, date is converted to a datetype
            if(dateType(date.value)==1):
                date=monthToDate(year,date.value)
            else:
                date=date.value.date()
            dates.append(date)
        else:
            dates.append(None)
       
    
    return categories, listdata, dates

#if dates given as months only, convert to dd/mm/year format
    #assumes first day of the month as day
def monthToDate(year,month):
    
    monthdict={'January':1, 'February':2, 'March':3, 'April':4, 'May':5, 'June':6, 'July':7, 
         'August':8, 'September':9, 'October':10, 'November':11, 'December':12}
    return d(year,monthdict[month],1)



#this function tells whether the format of the spreadsheet is in months, as we have seen
    #or if the format is mm/dd/yyyy
#1 for month format, 2 for mm/dd/yyyy
#0 for error
def dateType(value):
    
    #determine if a word is a month
    try:
        for month in months:
            if value.lower()==month.lower():
                return 1
    except AttributeError:
    #isolate the month if in dd/mm/year format
        return 2
    
    return 0

#this function finds empty categories or empty dates and removes 
#takes in the main data set, returns processed data set    
def filter(categories,data,dates): 
    print("removing any invalid data")
    #remove all -, --, and 'nan from the data.  This is used to represent nothing
    filterReport=''
    for i in range(0,len(data)):
        for j in range(0,len(data[0])):
            if(data[i][j]!=None):
                if ('-'== str(data[i][j]) or '--' == str(data[i][j]) or str(data[i][j])=='nan'):
                    data[i][j]=None
                try:
                    int(data[i][j])
                except:
                    data[i][j]=None
    #check if there is data in every category                
    i=0
    removedCount=0
    print("removing empty categories")
    for category in categories:
        remove=True
        for point in data[i]:
            if (point!=None):
                remove=False
            
        if(remove):
            filterReport+=("\nremoved category: "+categories[i]+" because no data was found")
            del categories[i]
            del data[i]
            removedCount+=1
            
        i+=1
    remove=True
    removedCount=0
    print("removing empty dates")
    #check if there is data in every date
    for i in range(0,len(dates)):
        remove=True
        
        for j in range(0,len(data)):
                
            
            if(data[j][i-removedCount]!=None):
                remove=False
        if dates[i-removedCount]==None:
            remove=True
        if(remove):
            
            filterReport+=("\nremoved date: "+str(dates[i-removedCount])+" because no data was found")
            
            del dates[i-removedCount]
            for j in range(0,len(data)):
                del data[j][i-removedCount]
            
            removedCount+=1
        
    print("done filtering")
    return (categories,data,dates,filterReport)

#this function will create a new spreadsheet populated with data from the given
#dataset
#filename tells where the file will save to
def toNewSpreadSheet(categories,data,dates,filename):
    wb = Workbook()
    dest_filename = filename
    ws1 = wb.active
    ws1.title = "Data spreadsheet"
    ws1['A1']='POLLEN DATA STORAGE'
    ws1=toWs(categories,data,dates,ws1)
    wb.save(filename = dest_filename)
    
    #used by toNewSpreadSheet and toMainSpreadSheet to properly format into
    #excel ws format
def toWs(categories,data,dates,ws1):
    
    for i in range(0,len(categories)):
        ws1.cell(row=1,column=i+2).value=categories[i]
    for i in range(0,len(dates)):
        ws1.cell(row=i+2,column=1).value=dates[i]
    for i in range(0, len(data)):
        for j in range(0,len(data[0])):
            ws1.cell(row=j+2,column=i+2).value=data[i][j]
    return ws1

#should only be used on the master spreadsheet's file
#loads from master spreadsheet to dataset simply
def loadData(fileName):
    try:
        ws1=readFile(fileName)
        categories,data,dates=toList(ws1, 1,False)
    except FileNotFoundError:
      categories,data,dates=[[],[],[]]  
    
    return (categories,data,dates)
def loadDataWithP(fileName):
    try:
        ws1=readFile(fileName)
        categories,data,dates=toList(ws1, 1,True)
    except FileNotFoundError:
      categories,data,dates=[[],[],[]]  
      
    return (categories,data,dates)
#brute force changing pollen category names, only used by the view/edit pollen
#categories button
def rewriteCategories(categories,fileName):
    data=loadDataWithP(fileName)
    
    toNewSpreadSheet(categories,data[1],data[2],fileName)
#merges new dataset into the main spreadsheet
#takes in dataset,saves file
#could be modified to return merged data set if needed
def toMainSpreadSheet(categories,datas,dates,filename):
    try:
        wb = load_workbook(filename)
        mainSpreadSheet=wb.active
        mainC,mainData,mainDates=toList(mainSpreadSheet,1,True)
    except FileNotFoundError:
        wb = Workbook()
        mainSpreadSheet=wb.active
        mainC,mainData,mainDates=[[],[],[]]
    #grab the active worksheet
    newCats=[]
    mergeReport=""
    #prepare dates for merging
    for i in range(0,len(datas)):
        
        mainDates,datas[i],mainData,assimilateReport=assimilate(datas[i],dates,mainDates,mainData)
        mergeReport+=assimilateReport
        #prepares categories for merging
    for i in range(0,len(categories)):
        add=True
        for j in range(0,len(mainC)): 
            #checks if there are matching categories
            minlen=min(len(categories[i]),len(mainC[j].split('(p')[0]))
            #if shortest category is less than 4 in length, compare literally
            #our collaborator uses abbreviations of 4 letters, anything below is
            #not an abbreviation. prevents merging of NO2, NO, NOx
            if(minlen<4):
                minlen=4
            if categories[i][0:minlen].lower()==mainC[j].split('(p')[0][0:minlen].lower():
                #special case for total pollen category so that
                #it is always replaced by newer data
                if categories[i][0:minlen].lower()=='total pollen':
                    mainData[j],updatedDataIndex=datas[i],[]
                else:
                    #proper array merge using helper
                    mainData[j],updatedDataIndex=mergeArray(datas[i],mainData[j])
                mergeReport+="\nMerging data in category "+categories[i]+" into "+mainC[j]
                
                for k in range(0,len(updatedDataIndex)):
                    pass
                    mergeReport+="\n Overwriting data in category "+mainC[j]+" at time "+str(mainDates[updatedDataIndex[k]])+" with "+str(datas[i][updatedDataIndex[k]]) 
                add=False
        if(add):
            #if there is no matching category, append it to the end
            mergeReport+= "\nCreating new Category: "+categories[i]
            #we use newcats so that the category dialogue can ask which are 
            #a pollen category
            newCats.append(categories[i])
            
            mainData.append(datas[i])
    #checking that we are not performing the total pollen merge from button,
    #asking user if this is a pollen category would be stupid and ignored        
    if len(newCats)>0 and not (newCats[0]=='Total Pollen' and len(newCats)==1):  
        #runs pollen category generator
        ex = pollenCategories(newCats)
        ex.show()
        if(ex.exec_()):
            for cat in ex.newCategories:
                mainC.append(cat)
    #if it is the total pollen merge from button, just append without asking   
    elif len(newCats)>0:
        mainC.append(newCats[0])
    #save the file
    mainSpreadSheet=toWs(mainC,mainData,mainDates,mainSpreadSheet)
    wb.save(filename)
    return (mergeReport)
#helper that merges two arrays of the same length together, returns result    
def mergeArray(arr1,arr2):
    updatedDataIndex=[]
    for i in range(0,len(arr1)):
        if(arr1[i]==arr2[i]):
            pass
        elif(arr1[i]==None):
            arr1[i]=arr2[i]
        else:
            updatedDataIndex.append(i)
    return arr1,updatedDataIndex

#helper function that properly sorts the new data into the main data
#adds new dates into the main date array if needed
#returns mainDates, newData(the newly formatted array that can be merged into mainData)
    #and main data with the new spaces for dates added
def assimilate(data,dates,mainDates,mainData):
    newData=[None]*len(mainDates)
    mergeReport=""
    
    for i in range(0,len(dates)):
        futureDate=True
        for j in range(0, len(mainDates)):
            
            if(dates[i]<mainDates[j]):
                mergeReport+='\nAdded Date '+str(dates[i])+" to file"
                mainDates.insert(j,dates[i])
                newData.insert(j,data[i])
                futureDate=False
                
                for k in range(0,len(mainData)):
                    mainData[k].insert(j,None)
                    
                    
                
                break
            if(dates[i]==mainDates[j]):
                
                newData[j]=data[i]
                futureDate=False
                break
        if(futureDate):
            mainDates.append(dates[i])
            newData.append(data[i])
            for k in range(0,len(mainData)):
                    mainData[k].append(None)
    
    return mainDates,newData,mainData,mergeReport
def calcTotalPollen(filename):
    dataToSum=[]
    
    categories,data,dates=loadDataWithP(filename)
    if len(categories)==0:
        
        raise FileNotFoundError
    sums=[[0]*len(dates)]
    pollenCats=0
    for i in range(len(categories)):
        if len(categories[i].split('(p'))==2:
            dataToSum.append(data[i])
            pollenCats+=1
            
    for j in range(len(dates)):
        isNone=True
        for i in range(0,pollenCats):
            if(dataToSum[i][j]!=None):
                sums[0][j]+=dataToSum[i][j]
                isNone=False
        if isNone:
            sums[0][j]=None
       
    
    report="Total Pollen per Day Calculated."
    report+=toMainSpreadSheet(["Total Pollen"],sums,dates,filename)
    return report

"""
#read in new file
ws=readFile('pollenData.xlsx')
print(findYear(ws))
print(findOrientation(ws))
datas = toList(ws,findOrientation(ws))
categories,data,dates,report=filter(datas[0],datas[1],datas[2])
print (report)

toNewSpreadSheet(categories,data,dates,'test.xlsx')

ws=readFile('exampleData.xlsx')
print(findYear(ws))
print(findOrientation(ws))
datas = toList(ws,findOrientation(ws))
categories,data,dates,report=filter(datas[0],datas[1],datas[2])
print (report)
report=toMainSpreadSheet(categories,data,dates,'test.xlsx')
print(report)
ws=readFile('Copy of DAILY TEMPLATE  STATION_ROCDPC_2019.xls')
print(findYear(ws))
print(findOrientation(ws))
datas = toList(ws,findOrientation(ws))
categories,data,dates,report=filter(datas[0],datas[1],datas[2])
print (report)
report=toMainSpreadSheet(categories,data,dates,'test.xlsx')
print(report)

print("done")

loadData('test.xlsx')
"""